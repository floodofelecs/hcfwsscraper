from re import split
from collections import defaultdict
from datetime import date
import csv
from re import sub
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import os
from openpyxl import load_workbook

#TO DO
# Q: Search and learn how to deploy this and make it run automatically at 1pm daily for example
# A: Run Every so often, e.g.: Once a day at noon or every 10 minutes: - just about every OS has 
# a scheduler that can be set to invoke specific tasks with configurable periodicity 
# - the best known one on Linux/Unix/OS-X is chron or chrontab.

#1. function that splits the text and returns the number and name of the sensor
#2. function that returns the sensor number given name
#3. function that returns sensor bayou/creek given the number
#4. Create a library that gets the total rain fall last 30 mins, 1 hour, day or between given two times



# def get_sensor_bayou(sens_dict, sen_name):
#     """
#     A function that takes a sensor name and the sensors' map and returns the sensor's number 
#     """
#     for key in sens_dict.keys():
#         if sens_dict[key] == sens


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a') # pylint: disable=abstract-class-instantiated

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# Create a web driver using chrome webdriver and send the link of the website to that driver
driver = webdriver.Chrome()


###########################
# Creating url            #
###########################

today = date.today()

d1 = today.strftime("%m/%d/%Y")

# print("Date:" + d1)
my_url = "https://www.harriscountyfws.org/GageDetail/Index/100?From=" + d1 + "%2012:00%20PM&span=24%20Hours&r=1&v=rainfall&selIdx=1"


# my_url = "https://www.harriscountyfws.org/GageDetail/Index/100?From=06/21/2021%2012:00%20PM&span=24%20Hours&r=1&v=rainfall&selIdx=1"



driver.get(my_url)

# Creates an implicit 10 second for any element to load 
# This allows us to avoid errors due to pages' loading 
driver.implicitly_wait(10)

####################
# Agency Selection #
####################

#Getting the agency drop down menu button and clicking it
agency_dropdown_button = driver.find_element_by_id("RegionComboBox_B-1")
agency_dropdown_button.click()


# Assigning sensors their locations according to the excel sheets published by the harris county flood warning system monthly reports
sens_num_loc = {
    "100":"Clear Creek", "110":"Clear Creek", "120":"Clear Creek","130":"Clear Creek", "135":"Clear Creek", "150":"Clear Creek","170":"Clear Creek",
    "175":"Clear Creek", "180":"Clear Creek", "190":"Clear Creek",
    "105":"Clear Creek Tributaries", "115":"Clear Creek Tributaries", "125":"Clear Creek Tributaries", "140":"Clear Creek Tributaries",
    "160":"Clear Creek Tributaries", "200":"Clear Creek Tributaries", "610":"Clear Creek Tributaries",
    "210":"Armand Bayou and Tributaries", "220":"Armand Bayou and Tributaries", "230":"Armand Bayou and Tributaries",
    "240":"Armand Bayou and Tributaries", "245":"Armand Bayou and Tributaries", "250":"Armand Bayou and Tributaries", "270":"Armand Bayou and Tributaries",
    "310":"Sims and Vince Bayou", "320":"Sims and Vince Bayou", "340":"Sims and Vince Bayou", "360":"Sims and Vince Bayou", "370":"Sims and Vince Bayou",
    "380":"Sims and Vince Bayou", "910":"Sims and Vince Bayou", "920":"Sims and Vince Bayou", "940":"Sims and Vince Bayou",
    "400":"Brays and Keegans Bayou", "405":"Brays and Keegans Bayou", "410":"Brays and Keegans Bayou", "420":"Brays and Keegans Bayou", 
    "430":"Brays and Keegans Bayou", "435":"Brays and Keegans Bayou", "440":"Brays and Keegans Bayou", "445":"Brays and Keegans Bayou", 
    "460":"Brays and Keegans Bayou", "465":"Brays and Keegans Bayou", "470":"Brays and Keegans Bayou", "475":"Brays and Keegans Bayou", 
    "485":"Brays and Keegans Bayou", "1020":"Brays and Keegans Bayou", 
    "480":"Keegans Bayou", "490":"Keegans Bayou", "495":"Keegans Bayou", 
    "510":"White Oak Bayou and Tributaries", "520":"White Oak Bayou and Tributaries", "530":"White Oak Bayou and Tributaries", 
    "535":"White Oak Bayou and Tributaries", "540":"White Oak Bayou and Tributaries", "545":"White Oak Bayou and Tributaries", 
    "550":"White Oak Bayou and Tributaries", "555":"White Oak Bayou and Tributaries", "560":"White Oak Bayou and Tributaries", 
    "570":"White Oak Bayou and Tributaries", "575":"White Oak Bayou and Tributaries", "580":"White Oak Bayou and Tributaries", 
    "582":"White Oak Bayou and Tributaries", "585":"White Oak Bayou and Tributaries", "590":"White Oak Bayou and Tributaries", 
    "595":"White Oak Bayou and Tributaries", "1000":"White Oak Bayou and Tributaries", 
    "605":"Cedar and Little Cedar Bayou and Goose Creek", "620":"Cedar and Little Cedar Bayou and Goose Creek", 
    "640":"Cedar and Little Cedar Bayou and Goose Creek", "650":"Cedar and Little Cedar Bayou and Goose Creek", 
    "660":"Cedar and Little Cedar Bayou and Goose Creek", "1520":"OutOfService", 
    "1540":"Cedar and Little Cedar Bayou and Goose Creek", "1720":"Cedar and Little Cedar Bayou and Goose Creek", 
    "1725":"Cedar and Little Cedar Bayou and Goose Creek", "1730":"Cedar and Little Cedar Bayou and Goose Creek", 
    "1740":"Cedar and Little Cedar Bayou and Goose Creek", "1745":"Cedar and Little Cedar Bayou and Goose Creek", 
    "710":"Luce and Jackson Bayou and San Jacinto", "720":"Luce and Jackson Bayou and San Jacinto", "740":"Luce and Jackson Bayou and San Jacinto", 
    "750":"Luce and Jackson Bayou and San Jacinto", "755":"Luce and Jackson Bayou and San Jacinto", "760":"Luce and Jackson Bayou and San Jacinto", 
    "765":"Luce and Jackson Bayou and San Jacinto", "770":"OutOfService", "780":"Luce and Jackson Bayou and San Jacinto", 
    "785":"Luce and Jackson Bayou and San Jacinto", "790":"Luce and Jackson Bayou and San Jacinto", "795":"Luce and Jackson Bayou and San Jacinto", 
    "1840":"Luce and Jackson Bayou and San Jacinto", "1930":"Luce and Jackson Bayou and San Jacinto", "1940":"Luce and Jackson Bayou and San Jacinto", 
    "1960":"Luce and Jackson Bayou and San Jacinto", "1975":"Luce and Jackson Bayou and San Jacinto", 
    "820":"Hunting and Carpenters Bayou", "830":"Hunting and Carpenters Bayou", "840":"Hunting and Carpenters Bayou", 
    "1420":"Hunting and Carpenters Bayou", "1440":"Hunting and Carpenters Bayou", "1460":"Hunting and Carpenters Bayou", 
    "1040":"Spring and Willow Creek", "1050":"Spring and Willow Creek", "1052":"Spring and Willow Creek", "1054":"Spring and Willow Creek", 
    "1055":"Spring and Willow Creek", "1056":"Spring and Willow Creek", "1060":"Spring and Willow Creek", "1070":"Spring and Willow Creek", 
    "1072":"Spring and Willow Creek", "1074":"Spring and Willow Creek", "1075":"Spring and Willow Creek", "1076":"Spring and Willow Creek", 
    "1080":"Spring and Willow Creek", "1084":"Spring and Willow Creek", "1086":"Spring and Willow Creek", "1090":"Spring and Willow Creek", 
    "1320":"Spring and Willow Creek", "1340":"Spring and Willow Creek",
    "1110":"Cypress Creek", "1115":"Cypress Creek", "1120":"Cypress Creek", "1130":"Cypress Creek", "1140":"Cypress Creek", 
    "1150":"Cypress Creek", "1160":"Cypress Creek", "1165":"Cypress Creek", "1170":"Cypress Creek", "1175":"Cypress Creek", 
    "1180":"Cypress Creek", "1185":"Cypress Creek", "1186":"Cypress Creek", "1190":"Cypress Creek", "1195":"Cypress Creek", 
    "1210":"Little Cypress Creek", "1220":"Little Cypress Creek", "1230":"Little Cypress Creek", 
    "1600":"Greens Bayou", "1610":"Greens Bayou", "1620":"Greens Bayou", "1640":"Greens Bayou", "1645":"Greens Bayou", "1660":"Greens Bayou", 
    "1665":"Greens Bayou", "1670":"Greens Bayou", "1685":"Greens Bayou", "1695":"Greens Bayou", 
    "1630":"Halls and Garners Bayou", "1650":"Halls and Garners Bayou", "1655":"Halls and Garners Bayou", "1675":"Halls and Garners Bayou", 
    "1680":"Halls and Garners Bayou", "1690":"Halls and Garners Bayou", 
    "2010":"Addicks and Barker Reservoir", "2015":"Addicks and Barker Reservoir", "2020":"Addicks and Barker Reservoir", "2025":"Addicks and Barker Reservoir", 
    "2030":"Addicks and Barker Reservoir", "2040":"Addicks and Barker Reservoir", "2050":"Addicks and Barker Reservoir", "2060":"Addicks and Barker Reservoir", 
    "2090":"Addicks and Barker Reservoir", "2110":"Addicks and Barker Reservoir", "2120":"Addicks and Barker Reservoir", "2130":"Addicks and Barker Reservoir", 
    "2140":"Addicks and Barker Reservoir", "2150":"Addicks and Barker Reservoir", "2160":"Addicks and Barker Reservoir", "2170":"Addicks and Barker Reservoir", 
    "2180":"Addicks and Barker Reservoir", "2190":"Addicks and Barker Reservoir", 
    "2210":"Buffalo Bayou", "2220":"Buffalo Bayou", "2240":"Buffalo Bayou", "2250":"Buffalo Bayou", "2253":"Buffalo Bayou", "2255":"Buffalo Bayou", 
    "2260":"Buffalo Bayou", "2265":"Buffalo Bayou", "2270":"Buffalo Bayou", "2280":"Buffalo Bayou", "2290":"Buffalo Bayou", 
}

sens_num_to_loc = defaultdict(int)

for key in sens_num_loc.keys():
    sens_num_to_loc[key] = sens_num_loc[key]


try:
    ###########################
    # Harris County Selection #
    ###########################

    # Setting an explicit wait of 10 seconds 
    wait = WebDriverWait(driver, 10)
    
    # Telling the program to wait as long as required (with a max of 10 seconds)
    # Until Harris County becomes clickable then store it in Harris County
    harris_county = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "RegionComboBox_DDD_L_LBI7T0"))
    )
    # Select harris county by clicking in the dropdown menu
    harris_county.click()
except:
    print("Harris County Selection failed")
    driver.quit()


try:
    #################################
    # Getting past 24 hours of data #
    #################################

    # Getting the last hours/days dropdown menu
    last_dropdown_button = driver.find_element_by_id("TimeFrameComboBox_B-1")
    last_dropdown_button.click()

    #Selecting the 24 hours option from the dropdown menu
    last_day = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "TimeFrameComboBox_DDD_L_LBI4T0"))
    )
    # Select harris county by clicking in the dropdown menu
    last_day.click()
except:
    print("Getting past 24 hours option failed")
    driver.quit()

try:
    #################################
    # Selecting diffreent locations #
    #################################

    # Click on the drop down button of location to select a new location
    loc_dropdown_button = driver.find_element_by_id("SiteComboBox_B-1")
    loc_dropdown_button.click()
 

    
    # f= open("/Users/mohamedabead/Desktop/vip/besmellah.txt","w")


    for i in range(1, 10):
        try:
            # Create the location ID using the pattern the locations were created with
            loc_id = "SiteComboBox_DDD_L_LBI" + str(i) + "T0"
            print("i: ", i, "  loc_id: ", loc_id)
            # Wait up to 10 seconds for the location to exist on the page
            # If the element is present, it clicks it
            location = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, loc_id))
            )
            # Wait up to 10 seconds until the location is clickable 
            # If the element is clickable, it gets it
            location = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, loc_id))
            )
            # Click on the element
            location.click()
        except:
                print("Selecting the location with id: " + loc_id + "failed")
                driver.quit()
        
        
        try:
            ####################################
            # Make Sure you're in rainfall tab #
            ####################################
            
            # Waiting explicitly 3 seconds because the tab crahses if the wait was implicit 
            # This could be improved
            # So far I tried implicitly and tried click on the li or the b elements instead of the a element 
            time.sleep(3)
            rainfall = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.LINK_TEXT, "Rainfall"))
            )

            rainfall = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.LINK_TEXT, "Rainfall"))
            )
            
            rainfall.click()
        except:
                print("Selecting the rainfall tab failed")
                driver.quit()


        try:
            ####################################
            # Getting data for every 5 minutes #
            ####################################
            data_intervals_button = wait.until(
                EC.element_to_be_clickable((By.ID, "IncrementComboBox_B-1"))
            )
            data_intervals_button.click()
            # Wait until (5 minutes) is present and clickable. Then click it
            five_minute_intervals = wait.until(
                EC.presence_of_element_located((By.ID, "IncrementComboBox_DDD_L_LBI0T0"))
            )

            five_minute_intervals = wait.until(
                EC.element_to_be_clickable((By.ID, "IncrementComboBox_DDD_L_LBI0T0"))
            )
            
            five_minute_intervals.click()
        except:
                print("Selecting the every 5 minute option failed")
                driver.quit()

        time.sleep(2)

        ######################################
        # Getting table data for each 5 mins #
        ######################################

        #Getting current page parsed html using beautifulsoup
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'lxml')
        
        #Getting the element of the selected location sensor 
        loc_name_e = soup.find(id=loc_id)

        #Get the name as string (e.g. "100: 100 Clear Lake 2nd Outlet @ SH 146")
        loc_name = loc_name_e.text.strip()

        # Split the location name to two strings "100", 
        split_loc_name = loc_name.split(":", 1)

        sensor_num = split_loc_name[0].strip()
        sensor_name = split_loc_name[1].strip().replace(sensor_num, '')
        sensor_name = sensor_name[:29]
        sensor_name_num = split_loc_name[1].strip()
        sensor_name_num = sensor_name_num[:29]
        
        
        print("Sensor name: " + sensor_name)
        # print(type(sensor_num))

        #Getting the entire table
        rain_level_table = soup.find("table", {"id": "IncrementGridView_DXMainTable"})

        # Table contains rows of tr tags 
        # Each row(tr tag) has 3 elements (td tags): "time from", "time to", and "the rain in inches".

        #We first get all rows
        tr_tags = rain_level_table.findChildren("tr")
        print("Hal hayshta8al?")
        # print(type(sens_num_to_loc))
        # print("Hal sens_num_to_loc[sensor_num] eshta8al?")
        # print("Sensor_num: " + "*" + sensor_num + "*")
        # print(sens_num_to_loc["145"])

        print("*" + (str) (sens_num_to_loc[sensor_num]) + "*")

        print("eshta8al")
        if sens_num_to_loc[sensor_num] != 0 and sens_num_to_loc[sensor_num] != "OutOfService":
            # Create path to be saved at using fielname of the bayou the sensor is 
            filename_loc = "/Users/mohamedabead/Desktop/vip/data/" +  sens_num_to_loc[sensor_num] + ".xlsx"

            #Creating the initials headers that will store the data
            dict_of_data =    {
                    "Reading Data from": [],
                    "Reading Data To": [],
                    "Rain in inches": [],
                }


            # f.write("\n\n\n\n" + sens_num_to_loc[sensor_num] + "\n \n\n\n ")

            #For each row (tr tag)
            for tr in tr_tags:
                #We get the row elements (children). These elements are td tags 
                td_tags = tr.findChildren("td")
                # We get the data in each cell in order 
                # We use the strip to remove all the spaces
                date_from = td_tags[0].text.strip()
                date_to = td_tags[1].text.strip()
                rain_in_inch = td_tags[2].text.strip()
                # If the data is there write it to the file
                if date_from != "" and date_from != "\n" and date_to != "" and date_to != "\n" and rain_in_inch != "" and rain_in_inch != "\n":
                    # Add data 
                    (dict_of_data["Reading Data from"]).append(date_from)
                    (dict_of_data["Reading Data To"]).append(date_to)
                    (dict_of_data["Rain in inches"]).append(rain_in_inch)


                    # f.write("Date from: " + date_from + "\n")
                    # f.write("Date to: " + date_to + "\n")
                    # f.write("Rain inches: " + rain_in_inch + "\n")
                    # f.write("\n *********************\n ")
            
            #Creating panda data frame
            panda_df = pd.DataFrame(dict_of_data)

            # Adding fetched data to the excel data 
            # panda_df.to_excel(filename_loc, sheet_name = sensor_name, index=True)
            append_df_to_excel(filename_loc, panda_df, sheet_name=sensor_name_num, index=False)

            # f.write("\n\n\n\n ######## Ya wa3dy ###### \n \n\n\n ")


        
        if i != 199:
            loc_dropdown_button = driver.find_element_by_id("SiteComboBox_B-1")
            loc_dropdown_button.click()
    
# If there is an error, quits the driver
except:
    print("ERRROR in the fetching and saving data part")
    driver.quit()

time.sleep(3)
# f.close()
driver.quit()

# betwwen 770 or 780 out of service -> have rain data
# 1000, 1075, 1076, 1520, 1720, 1930, 2000, 2010, 2110 -> Have rain data 
